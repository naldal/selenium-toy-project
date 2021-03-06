from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
import time

cd = "chromedriver"
driver = webdriver.Chrome(cd)
driver.get("https://www.youtube.com")

keyword = ['가을', '네일아트']
chr_array = ['ㄱ', 'ㄴ', 'ㄷ', 'ㄹ', 'ㅁ', 'ㅂ', 'ㅅ', 'ㅇ', 'ㅈ', 'ㅊ', 'ㅋ', 'ㅌ', 'ㅍ', 'ㅎ']

search = driver.find_element_by_xpath('//*[@id="search"]')

wb = Workbook()

for sheet in keyword:
    ws = wb.create_sheet(sheet)
    column = 1
    search.send_keys(sheet + ' ?')

    for n in chr_array:
        row = 1
        ws.cell(row, column, n)
        search.send_keys(Keys.BACK_SPACE)
        search.send_keys(n)
        time.sleep(2)
        search.send_keys(Keys.ARROW_DOWN)
        time.sleep(2)

        html = driver.find_element_by_xpath('//*').get_attribute('outerHTML')
        soup = bs(html, 'html.parser')
        res = soup.find_all('div', class_='sbqs_c')

        for i in res:
            row += 1
            ws.cell(row, column, i.contents[1].contents[0])
        column += 1

    search.send_keys(Keys.CONTROL+'a')
    time.sleep(1)
    search.send_keys(Keys.DELETE)
    time.sleep(1)

wb.save('test.xlsx')
driver.close()
